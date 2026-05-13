#!/usr/bin/env python3
import json, base64, io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── HELPERS ───────────────────────────────────────────────────────────────────
def fill(c): return PatternFill('solid', fgColor=c.replace('#','').lstrip('0')[:6].zfill(6))
def thin(): return Side(style='thin', color='CCCCCC')
def bdr(): return Border(left=thin(),right=thin(),top=thin(),bottom=thin())
def aln(h='left',v='center',wrap=False): return Alignment(horizontal=h,vertical=v,wrap_text=wrap)
def w(ws,row,col,val,bold=False,fg='1E2A38',bg=None,align='left',size=10,border=True,wrap=False,italic=False,fmt=None,size_row=None):
    c=ws.cell(row=row,column=col,value=val)
    c.font=Font(name='Arial',bold=bold,color=fg,size=size,italic=italic)
    if bg: c.fill=fill(bg)
    c.alignment=aln(align,wrap=wrap)
    if border: c.border=bdr()
    if fmt: c.number_format=fmt
    return c

def get_status(tipo, costo):
    if not costo or costo<=0: return ''
    t=(tipo or '').lower()
    if 'purchase' in t or 'convers' in t or 'venta' in t or 'compra' in t:
        if costo<6000:  return '🏆 GANADOR'
        if costo<10000: return '✅ OK'
        if costo<18000: return '🔍 REVISAR'
        return '⛔ PAUSAR'
    elif 'lead' in t or 'formulario' in t:
        if costo<50:  return '🏆 GANADOR'
        if costo<90:  return '✅ OK'
        if costo<150: return '🔍 REVISAR'
        return '⛔ PAUSAR'
    else:
        if costo<20: return '🏆 GANADOR'
        if costo<40: return '✅ OK'
        if costo<80: return '🔍 REVISAR'
        return '⛔ PAUSAR'

def status_colors(status):
    if 'GANADOR' in status: return ('27AE60','E8F8F1')
    if 'PAUSAR'  in status: return ('C0392B','FDEDEC')
    if 'REVISAR' in status: return ('3730A3','EEF2FF')
    if 'OK'      in status: return ('1A5276','E8F4F8')
    return ('1E2A38','FFFFFF')

def fmt_pct(v):
    if v is None: return ''
    return ('+' if v>0 else '')+f'{v:.0f}%'

# ── LEER JSON ─────────────────────────────────────────────────────────────────
with open('/tmp/excel_data.json','r') as f:
    data = json.load(f)

label_actual   = data.get('label_actual','Mes actual')
label_anterior = data.get('label_anterior','Mes anterior')
marcas_list    = data.get('marcas',[])

wb = Workbook()
ws_res = wb.active
ws_res.title = 'RESULTADOS'

# ── HOJA RESULTADOS: HEADER ───────────────────────────────────────────────────
ws_res.merge_cells('B1:S1')
w(ws_res,1,2,f'INFORME META ADS — {label_actual.upper()}',bold=True,fg='FFFFFF',bg='1A1A2E',align='center',size=14,border=False)
ws_res.row_dimensions[1].height=36

ws_res.merge_cells('B2:S2')
w(ws_res,2,2,f'{label_actual}  vs  {label_anterior}  ·  Fuente: Meta Ads API  ·  Agrupado por campaña',fg='AAAAAA',bg='1A1A2E',align='center',size=10,border=False)
ws_res.row_dimensions[2].height=20

# Benchmarks fila 5
bench={13:'>1%',14:'',15:'>25%',16:'>12%',17:'>8%',18:'>3%',19:'>2x'}
for col,lbl in bench.items():
    w(ws_res,5,col,lbl,fg='888888',align='center',size=9,border=False,bg='F8F9FA')

# Headers fila 7
hdrs=['CAMPAÑA / ANUNCIO','Alcance','Impresiones','Frecuencia','Resultados',
      f'Resultado ant.\n({label_anterior})','%','Gasto (CLP)','Costo/Result.',
      'CTR%','CPM','Hook Rate%','Hold Rate 50%','Hold Rate 75%','Hold Rate 100%',
      'ROAS','Clicks sitio','STATUS']
cws=[35,12,13,10,12,16,8,14,12,8,10,10,13,13,13,8,11,14]
for i,h in enumerate(hdrs):
    w(ws_res,7,i+2,h,bold=True,fg='FFFFFF',bg='1A1A2E',align='center',size=10,wrap=True)
    ws_res.column_dimensions[get_column_letter(i+2)].width=cws[i]
ws_res.row_dimensions[7].height=45

rr=9  # fila actual en RESULTADOS

# ── HOJAS POR MARCA + RESULTADOS ──────────────────────────────────────────────
for marca_data in marcas_list:
    nombre_marca  = marca_data.get('marca_nombre', marca_data.get('nombre','Marca'))
    ads_actual    = marca_data.get('ads_actual',[])
    ads_anterior  = marca_data.get('ads_anterior',[])
    ant_map       = {a.get('anuncio',''): a for a in ads_anterior}

    # Totales de marca para header de hoja individual
    gasto_act_tot  = sum(float(a.get('gasto',0) or 0) for a in ads_actual)
    gasto_ant_tot  = sum(float(ant_map.get(a.get('anuncio',''),{}).get('gasto',0) or 0) for a in ads_actual)
    delta_gasto    = ((gasto_act_tot-gasto_ant_tot)/gasto_ant_tot*100) if gasto_ant_tot>0 else None
    n_anuncios     = len(ads_actual)

    # ── Fila de marca en RESULTADOS ──────────────────────────────────────────
    ws_res.merge_cells(f'B{rr}:S{rr}')
    w(ws_res,rr,2,f'  {nombre_marca.upper()}',bold=True,fg='FFFFFF',bg='1A1A2E',align='left',size=11)
    ws_res.row_dimensions[rr].height=26
    rr+=1

    # ── Crear hoja por marca ──────────────────────────────────────────────────
    ws_m = wb.create_sheet(title=nombre_marca[:31])

    ws_m.merge_cells('B1:R1')
    w(ws_m,1,2,f'{nombre_marca.upper()} — {label_actual.upper()} vs {label_anterior.upper()}',
      bold=True,fg='FFFFFF',bg='1A1A2E',align='center',size=13,border=False)
    ws_m.row_dimensions[1].height=32

    ws_m.merge_cells('B2:R2')
    gasto_str=f'Gasto actual: ${int(gasto_act_tot):,}  ·  Anterior: ${int(gasto_ant_tot):,}  ·  Δ Gasto: {fmt_pct(delta_gasto)}  ·  {n_anuncios} anuncios activos'
    w(ws_m,2,2,gasto_str,fg='AAAAAA',bg='16213E',align='center',size=10,border=False)
    ws_m.row_dimensions[2].height=18

    # Headers hoja marca (fila 4)
    hdrs_m=['ANUNCIO','OBJETIVO','GASTO ACT','GASTO ANT','Δ%','RES ACT','RES ANT','Δ%','CTR%','CPM','HOOK%','HOLD50%','HOLD75%','HOLD100%','ROAS','CLICKS','STATUS']
    cws_m=[35,14,12,12,8,12,12,8,8,10,10,10,10,10,8,10,14]
    for i,h in enumerate(hdrs_m):
        w(ws_m,4,i+2,h,bold=True,fg='FFFFFF',bg='2E4057',align='center',size=10,wrap=True)
        ws_m.column_dimensions[get_column_letter(i+2)].width=cws_m[i]
    ws_m.row_dimensions[4].height=40

    rm=6  # fila actual en hoja marca

    # ── Agrupar por campaña ───────────────────────────────────────────────────
    camps={}
    for ad in ads_actual:
        cn=ad.get('campaña','Sin campaña')
        if cn not in camps: camps[cn]=[]
        camps[cn].append(ad)

    for camp_name, ads in camps.items():
        # Header campaña en RESULTADOS
        ws_res.merge_cells(f'B{rr}:S{rr}')
        w(ws_res,rr,2,f'  ▸  {camp_name}',bold=True,fg='FFFFFF',bg='2E4057',align='left',size=10)
        ws_res.row_dimensions[rr].height=22
        rr+=1

        # Header campaña en hoja marca
        ws_m.merge_cells(f'B{rm}:R{rm}')
        w(ws_m,rm,2,f'  ▸  {camp_name}',bold=True,fg='FFFFFF',bg='2E4057',align='left',size=10)
        ws_m.row_dimensions[rm].height=22
        rm+=1

        tot_res=tot_res_ant=tot_gasto=tot_gasto_ant=0

        for ad in ads:
            nombre   = ad.get('anuncio','')
            tipo     = ad.get('tipo','')
            res      = float(ad.get('resultados',0) or 0)
            gasto    = float(ad.get('gasto',0) or 0)
            costo    = float(ad.get('costo',0) or 0)
            ctr      = float(ad.get('ctr',0) or 0)
            alcance  = float(ad.get('alcance',0) or 0)
            impr     = float(ad.get('impresiones',0) or 0)
            freq     = float(ad.get('frecuencia',0) or 0)
            hook     = float(ad.get('hook',0) or 0)
            hold50   = float(ad.get('hold50',0) or 0)
            hold75   = float(ad.get('hold75',0) or 0)
            hold100  = float(ad.get('hold100',0) or 0)
            roas     = float(ad.get('roas',0) or 0)
            clicks   = float(ad.get('clicks_sitio',0) or 0)
            cpm      = float(ad.get('cpm',0) or 0)

            ad_ant   = ant_map.get(nombre,{})
            res_ant  = float(ad_ant.get('resultados',0) or 0)
            gasto_ant= float(ad_ant.get('gasto',0) or 0)
            pct_res  = ((res-res_ant)/res_ant*100) if res_ant>0 else None
            pct_gasto= ((gasto-gasto_ant)/gasto_ant*100) if gasto_ant>0 else None
            status   = get_status(tipo, costo)
            st_fg, st_bg = status_colors(status)

            # Color fila por status
            bg_row='FFFFFF'
            if 'PAUSAR'  in status: bg_row='FDEDEC'
            elif 'GANADOR' in status: bg_row='EAFAF1'
            elif 'REVISAR' in status: bg_row='EEF2FF'

            # ── RESULTADOS sheet ──────────────────────────────────────────────
            w(ws_res,rr,2,nombre,bg=bg_row)
            w(ws_res,rr,3,int(alcance) if alcance else '',bg=bg_row,align='right')
            w(ws_res,rr,4,int(impr) if impr else '',bg=bg_row,align='right')
            w(ws_res,rr,5,round(freq,1) if freq else '',bg=bg_row,align='right')
            w(ws_res,rr,6,int(res) if res else '',bold=True,bg=bg_row,align='right')
            w(ws_res,rr,7,int(res_ant) if res_ant else '',italic=True,fg='888888',bg=bg_row,align='right')
            w(ws_res,rr,8,fmt_pct(pct_res) if pct_res is not None else '',bold=True,
              fg='27AE60' if (pct_res or 0)>0 else 'C0392B',bg=bg_row,align='center')
            w(ws_res,rr,9,int(gasto) if gasto else '',bg=bg_row,align='right',fmt='#,##0')
            w(ws_res,rr,10,int(costo) if costo else '',bg=bg_row,align='right',fmt='#,##0')
            w(ws_res,rr,11,round(ctr,2) if ctr else '',bg=bg_row,align='right')
            w(ws_res,rr,12,int(cpm) if cpm else '',bg=bg_row,align='right')
            w(ws_res,rr,13,'',bg=bg_row)  # hook col 13 vacía (benchmark label row)
            w(ws_res,rr,14,round(hook,1) if hook else '',bold=hook>=25,
              fg='27AE60' if hook>=25 else ('C0392B' if 0<hook<15 else '1E2A38'),bg=bg_row,align='right')
            w(ws_res,rr,15,round(hold50,1) if hold50 else '',bg=bg_row,align='right')
            w(ws_res,rr,16,round(hold75,1) if hold75 else '',bg=bg_row,align='right')
            w(ws_res,rr,17,round(hold100,1) if hold100 else '',bg=bg_row,align='right')
            w(ws_res,rr,18,round(roas,2) if roas else '',bg=bg_row,align='right')
            w(ws_res,rr,19,int(clicks) if clicks else '',bg=bg_row,align='right')
            w(ws_res,rr,20,status,bold=True,fg=st_fg,bg=st_bg,align='center') if False else None
            # STATUS en col 19... espera, veamos: hdrs tiene 18 items, empieza en col 2 → STATUS en col 19
            ws_res.cell(rr,19).value=None  # limpiar
            c_st=ws_res.cell(rr,19); c_st.value=''  # placeholder
            # Reescribir correctamente
            ws_res.cell(row=rr,column=19).value=''  
            # El STATUS va en columna 19 = col S
            cell=ws_res.cell(row=rr,column=19)
            cell.value=status
            cell.font=Font(name='Arial',bold=True,color=st_fg,size=10)
            cell.fill=fill(st_bg)
            cell.alignment=aln('center')
            cell.border=bdr()
            ws_res.row_dimensions[rr].height=18

            # ── MARCA sheet ───────────────────────────────────────────────────
            w(ws_m,rm,2,nombre,bg=bg_row)
            w(ws_m,rm,3,tipo,fg='888888',bg=bg_row,size=9)
            w(ws_m,rm,4,int(gasto) if gasto else '',bg=bg_row,align='right',fmt='#,##0')
            w(ws_m,rm,5,int(gasto_ant) if gasto_ant else '',italic=True,fg='888888',bg=bg_row,align='right',fmt='#,##0')
            w(ws_m,rm,6,fmt_pct(pct_gasto) if pct_gasto is not None else '',bold=True,
              fg='C0392B' if (pct_gasto or 0)>0 else '27AE60',bg=bg_row,align='center')
            w(ws_m,rm,7,int(res) if res else '',bold=True,bg=bg_row,align='right')
            w(ws_m,rm,8,int(res_ant) if res_ant else '',italic=True,fg='888888',bg=bg_row,align='right')
            w(ws_m,rm,9,fmt_pct(pct_res) if pct_res is not None else '',bold=True,
              fg='27AE60' if (pct_res or 0)>0 else 'C0392B',bg=bg_row,align='center')
            w(ws_m,rm,10,round(ctr,2) if ctr else '',bg=bg_row,align='right')
            w(ws_m,rm,11,int(cpm) if cpm else '',bg=bg_row,align='right')
            w(ws_m,rm,12,round(hook,1) if hook else '',bold=hook>=25,
              fg='27AE60' if hook>=25 else ('C0392B' if 0<hook<15 else '1E2A38'),bg=bg_row,align='right')
            w(ws_m,rm,13,round(hold50,1) if hold50 else '',bg=bg_row,align='right')
            w(ws_m,rm,14,round(hold75,1) if hold75 else '',bg=bg_row,align='right')
            w(ws_m,rm,15,round(hold100,1) if hold100 else '',bg=bg_row,align='right')
            w(ws_m,rm,16,round(roas,2) if roas else '',bg=bg_row,align='right')
            w(ws_m,rm,17,int(clicks) if clicks else '',bg=bg_row,align='right')
            cell_m=ws_m.cell(row=rm,column=18)
            cell_m.value=status
            cell_m.font=Font(name='Arial',bold=True,color=st_fg,size=10)
            cell_m.fill=fill(st_bg)
            cell_m.alignment=aln('center')
            cell_m.border=bdr()
            ws_m.row_dimensions[rm].height=18

            tot_res+=res; tot_res_ant+=res_ant
            tot_gasto+=gasto; tot_gasto_ant+=gasto_ant
            rr+=1; rm+=1

        # SUBTOTAL campaña en RESULTADOS
        ws_res.merge_cells(f'B{rr}:E{rr}')
        w(ws_res,rr,2,f'SUBTOTAL  {camp_name}',bold=True,fg='FFFFFF',bg='2E4057')
        w(ws_res,rr,6,int(tot_res),bold=True,fg='FFFFFF',bg='2E4057',align='right')
        w(ws_res,rr,7,int(tot_res_ant) if tot_res_ant else '',italic=True,fg='FFFFFF',bg='2E4057',align='right')
        if tot_res_ant>0:
            p=((tot_res-tot_res_ant)/tot_res_ant*100)
            w(ws_res,rr,8,fmt_pct(p),bold=True,fg='FFFFFF',bg='2E4057',align='center')
        w(ws_res,rr,9,int(tot_gasto),bold=True,fg='FFFFFF',bg='2E4057',align='right',fmt='#,##0')
        ws_res.row_dimensions[rr].height=20
        rr+=1

        # SUBTOTAL campaña en hoja marca
        ws_m.merge_cells(f'B{rm}:C{rm}')
        w(ws_m,rm,2,f'SUBTOTAL  {camp_name}',bold=True,fg='FFFFFF',bg='2E4057')
        w(ws_m,rm,4,int(tot_gasto),bold=True,fg='FFFFFF',bg='2E4057',align='right',fmt='#,##0')
        w(ws_m,rm,5,int(tot_gasto_ant) if tot_gasto_ant else '',italic=True,fg='FFFFFF',bg='2E4057',align='right',fmt='#,##0')
        if tot_gasto_ant>0:
            pg=((tot_gasto-tot_gasto_ant)/tot_gasto_ant*100)
            w(ws_m,rm,6,fmt_pct(pg),bold=True,fg='FFFFFF',bg='2E4057',align='center')
        w(ws_m,rm,7,int(tot_res),bold=True,fg='FFFFFF',bg='2E4057',align='right')
        w(ws_m,rm,8,int(tot_res_ant) if tot_res_ant else '',italic=True,fg='FFFFFF',bg='2E4057',align='right')
        if tot_res_ant>0:
            pr=((tot_res-tot_res_ant)/tot_res_ant*100)
            w(ws_m,rm,9,fmt_pct(pr),bold=True,fg='FFFFFF',bg='2E4057',align='center')
        ws_m.row_dimensions[rm].height=20
        rm+=2

    rr+=1  # espacio entre marcas
    ws_m.freeze_panes='B5'
    ws_m.sheet_view.zoomScale=90

ws_res.freeze_panes='B8'
ws_res.sheet_view.zoomScale=90

buf=io.BytesIO()
wb.save(buf)
buf.seek(0)
print(base64.b64encode(buf.read()).decode('utf-8'))
